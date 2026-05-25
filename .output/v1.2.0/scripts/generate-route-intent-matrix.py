#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""v1.2.0 全量路由-指令穷举清单"""

from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_OUT = "全量路由-指令穷举清单-v1.2.0.xlsx"
OUT = Path(__file__).resolve().parents[1] / (sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT)
HEADERS = [
    "功能名称", "当前步骤", "目标步骤", "目标子步骤/界面名称",
    "交互模式", "指令示例", "已提供", "待提供",
]

COMPLETE_ROWS = []
PARTIAL_ROWS = []
CLICK_ROWS = []
FILL_ROWS = []
MISC_ROWS = []

FI_PLAN, FI_QUOTE, FI_ORDER, FI_MISC = 0, 1, 2, 3

PLAN_STEP_ORDER = {
    "对话页": 0, "选客户": 0, "入口": 1, "查看历史": 2, "需求录入": 3, "选品": 4,
    "方案预览": 5, "方案模板": 6, "方案成果": 7,
}
QUOTE_STEP_ORDER = {
    "对话页": 0, "选客户": 0, "入口": 1, "报价来源": 2, "选择方案": 3, "需求录入": 4,
    "选品": 5, "逐项报价": 6, "报价单模板": 7, "报价成果": 8,
}
ORDER_STEP_ORDER = {
    "对话页": 0, "选客户": 0, "入口": 1, "下单来源": 2, "选择报价单": 3, "选品": 4,
    "逐项报价": 5, "订单确认": 6, "订单成功": 7,
}
BRANCH_ORDER = {"按方案": 0, "自选": 1, "按报价单": 0}
CUST_ORDER = {"新客户": 0, "老客户": 1}
QUOTE_BRANCHES = [("按方案", "新客户"), ("按方案", "老客户"), ("自选", "新客户"), ("自选", "老客户")]
ORDER_BRANCHES = [("按报价单", "新客户"), ("按报价单", "老客户"), ("自选", "新客户"), ("自选", "老客户")]

PLAN_ENTRY_MISSING = "查看历史数据 或 创建新方案（二选一）"
QUOTE_ENTRY_MISSING = "查看历史报价单 或 新建报价（二选一）"


def plan_feat(customer):
    return f"方案速配（{customer}）"


def quote_feat(branch, customer):
    return f"产品报价（{branch}·{customer}）"


def order_feat(branch, customer):
    return f"确认下单（{branch}·{customer}）"


def pkey(fi, step, smap, branch=None, customer=None, sub=0):
    b = BRANCH_ORDER.get(branch, -1) if branch else -1
    c = CUST_ORDER.get(customer, -1) if customer else -1
    return (fi, smap.get(step, 99), b, c, sub)


def row(feature, cur, tgt, screen, mode, example, provided, missing="—"):
    return (feature, cur, tgt, screen, mode, example, provided, missing)


def add_complete(feature, cur, tgt, screen, example, provided):
    COMPLETE_ROWS.append(row(feature, cur, tgt, screen, "完整需求", example, provided))


def add_partial(sk, feature, cur, tgt, screen, example, provided, missing):
    PARTIAL_ROWS.append((sk, row(feature, cur, tgt, screen, "部分需求", example, provided, missing)))


def add_click(sk, feature, cur, tgt, screen, example):
    CLICK_ROWS.append((sk, row(feature, cur, tgt, screen, "模拟点击", example, "—", "—")))


def add_fill(sk, feature, cur, tgt, screen, example, provided, missing):
    FILL_ROWS.append((sk, row(feature, cur, tgt, screen, "表单填充", example, provided, missing)))


# ── 完整需求：对话页一句直达成果 ───────────────────────────
add_complete(plan_feat("新客户"), "对话页", "方案成果", "方案成果卡",
             "给深圳创源配伺服电机和传动齿轮箱各2台，用标准技术方案",
             "客户、需求描述、品名、数量、方案模板")
add_complete(plan_feat("老客户"), "对话页", "方案成果", "方案成果卡",
             "给华东精密配精密轴承组件A型3套、传动齿轮箱M3各2台，投标方案简版",
             "客户、品名、数量、方案模板")
for br, cu, ex, prov in [
    ("按方案", "新客户", "给深圳创源按方案报价，标准技术方案…标准销售报价单", "客户、方案、本单报价、报价单模板"),
    ("按方案", "老客户", "给华东精密按方案报价，标准技术方案，伺服电机4200…", "客户、方案、各行本单报价、报价单模板"),
    ("自选", "新客户", "给深圳创源报伺服电机10台单价2180打九折，标准销售报价单", "客户、品名、数量、单价/折扣、报价单模板"),
    ("自选", "老客户", "给华东精密报传动齿轮箱2台单价3680，标准销售报价单", "客户、品名、数量、本单报价、报价单模板"),
]:
    add_complete(quote_feat(br, cu), "对话页", "报价成果", "报价单成果卡", ex, prov)
for br, cu, ex, prov in [
    ("按报价单", "新客户", "给深圳创源按报价单下单", "客户、报价单标识"),
    ("按报价单", "老客户", "给华东精密按报价单下单，报价单编号", "客户、报价单标识"),
    ("自选", "新客户", "给深圳创源下单，配伺服电机5台单价2180", "客户、品名、数量、本单报价"),
    ("自选", "老客户", "给华东精密下单，配伺服电机10台单价2100", "客户、品名、数量、本单报价"),
]:
    add_complete(order_feat(br, cu), "对话页", "订单成功", "订单提交成功卡", ex, prov)

# ── 部分需求 · 方案速配 ───────────────────────────────────
for cu in ("新客户", "老客户"):
    add_partial(pkey(FI_PLAN, "入口", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "对话页", "入口", "方案速配入口卡", "配个方案（已选客户）",
                "客户、功能意图", PLAN_ENTRY_MISSING)
    add_partial(pkey(FI_PLAN, "入口", PLAN_STEP_ORDER, customer=cu, sub=1),
                plan_feat(cu), "入口", "入口", "方案速配入口卡", "筛选 伺服（仍停在入口卡）",
                "客户", PLAN_ENTRY_MISSING)
add_partial(pkey(FI_PLAN, "入口", PLAN_STEP_ORDER, customer="新客户"),
            plan_feat("新客户"), "对话页", "选客户", "选客户消息卡", "配个方案（未选客户）",
            "功能意图", "客户")
add_partial(pkey(FI_PLAN, "查看历史", PLAN_STEP_ORDER, customer="老客户"),
            plan_feat("老客户"), "对话页", "查看历史", "历史方案列表卡", "查看历史方案",
            "客户、功能意图", "方案序号（多条时）")
for cu in ("新客户", "老客户"):
    demo = "深圳创源" if cu == "新客户" else "华东精密"
    tgt_demand = "需求录入" if cu == "新客户" else "选品"
    add_partial(pkey(FI_PLAN, "需求录入", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "对话页", tgt_demand, "需求描述卡", f"给{demo}配个方案",
                "客户、功能意图",
                "需求描述、选品" if cu == "新客户" else "选品（需求可选）")
    add_partial(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "选品", "选品", "方案选品卡", "预览方案（未勾选）",
                "客户" + ("、需求" if cu == "新客户" else ""), "选品勾选")
    add_partial(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "选品", "选品", "方案选品卡", f"给{demo}配液压泵站",
                "客户、部分品名", "可匹配选品")
    add_partial(pkey(FI_PLAN, "方案预览", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "方案预览", "方案模板", "方案预览卡", "生成方案（口语，未选模板）",
                "选品勾选", "方案模板")
    add_partial(pkey(FI_PLAN, "方案模板", PLAN_STEP_ORDER, customer=cu),
                plan_feat(cu), "方案模板", "方案模板", "方案模板消息卡", "保存方案（未选模板）",
                "选品、规格、数量", "方案模板")
    add_partial(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=1),
                plan_feat(cu), "选品", "选品", "方案选品卡", "加购（选品卡未勾选）",
                "客户、功能意图、已进选品卡" + ("、需求描述" if cu == "新客户" else ""),
                "选品勾选（至少勾选一种产品）")

# ── 部分需求 · 产品报价 ───────────────────────────────────
for br, cu in QUOTE_BRANCHES:
    sk = lambda st, s=0: pkey(FI_QUOTE, st, QUOTE_STEP_ORDER, br, cu, s)
    feat = quote_feat(br, cu)
    add_partial(sk("入口"), feat, "对话页", "入口", "产品报价入口卡", "报价（已选客户）",
                "客户、功能意图", QUOTE_ENTRY_MISSING)
    add_partial(sk("入口", 1), feat, "入口", "入口", "产品报价入口卡",
                "按方案报价（仍停在入口卡）" if br == "按方案" else "筛选 电机（仍停在入口卡）",
                "客户", QUOTE_ENTRY_MISSING)
    add_partial(sk("报价来源"), feat, "对话页", "报价来源", "报价来源卡", "报价",
                "客户、功能意图", "按方案报价 或 直接选品报价（二选一）")
    if br == "按方案":
        add_partial(sk("选择方案"), feat, "报价来源", "选择方案", "选择方案卡", "按方案报价",
                    "客户、来源=按方案", "方案名称/编号")
        add_partial(sk("选择方案", 1), feat, "选择方案", "选择方案", "选择方案卡",
                    "按方案报价 不存在的方案名", "客户", "有效方案")
        if cu == "新客户":
            add_partial(sk("报价来源", 1), feat, "报价来源", "报价来源", "报价来源卡",
                        "按方案报价（无方案）", "客户", "本客户方案")
    if br == "自选":
        add_partial(sk("选品"), feat, "选品", "选品", "报价选品卡",
                    "给深圳创源报价 无效品" if cu == "新客户" else "下一步（未选品）",
                    "客户" + ("、品名" if cu == "新客户" else ""),
                    "可匹配产品" if cu == "新客户" else "已选产品")
    add_partial(sk("逐项报价"), feat, "逐项报价", "逐项报价", "逐项报价消息卡", "选择模板（未填价）",
                "选品明细", "每项本单报价")
    add_partial(sk("报价单模板"), feat, "报价单模板", "报价单模板", "报价单模板消息卡",
                "生成报价单（未选模板）", "明细", "报价单模板")

# ── 部分需求 · 确认下单 ───────────────────────────────────
for br, cu in ORDER_BRANCHES:
    sk = lambda st, s=0: pkey(FI_ORDER, st, ORDER_STEP_ORDER, br, cu, s)
    feat = order_feat(br, cu)
    add_partial(sk("入口"), feat, "对话页", "选客户", "确认下单入口", "确认下单 / 生成订单",
                "功能意图", "客户")
    add_partial(sk("下单来源"), feat, "对话页", "下单来源", "下单来源卡", "生成订单（已选客户）",
                "客户、功能意图", "按报价单下单 或 直接选品（二选一）")
    if br == "按报价单":
        if cu == "老客户":
            add_partial(sk("下单来源", 1), feat, "下单来源", "下单来源", "下单来源卡",
                        "按报价单下单（无报价单）", "客户", "本客户报价单")
        add_partial(sk("选择报价单"), feat, "下单来源", "选择报价单", "选择报价单卡", "按报价单下单",
                    "客户、来源=按报价单", "报价单编号/模板")
        add_partial(sk("选择报价单", 1), feat, "选择报价单", "选择报价单", "选择报价单卡",
                    "按报价单下单 不存在单号", "客户", "有效报价单")
    if br == "自选":
        add_partial(sk("选品"), feat, "选品", "选品", "订单选品卡", "生成订单（未选品）",
                    "客户", "已选产品")
    add_partial(sk("逐项报价"), feat, "逐项报价", "逐项报价", "逐项报价消息卡", "确认下单（未填价）",
                "选品明细", "每项本单报价")
    add_partial(sk("订单确认"), feat, "订单确认", "订单确认", "订单确认消息卡", "确认下单（无明细）",
                "—", "订单来源与明细")

add_partial((FI_MISC, 0, -1, -1, 0), "对话页", "对话页", "选客户", "选客户消息卡",
            "配个方案 / 报价 / 生成订单", "功能意图", "客户")
add_partial((FI_MISC, 1, -1, -1, 0), "对话页", "对话页", "对话页", "技能条", "深圳创源",
            "客户", "功能意图")
add_partial((FI_MISC, 2, -1, -1, 0), plan_feat("老客户"), "选品", "跨功能切换",
            "方案选品卡", "方案选品卡中说「报价」", "客户、方案上下文", "是否带入报价流程")

# ── 模拟点击 ─────────────────────────────────────────────
for cu in ("新客户", "老客户"):
    add_click(pkey(FI_PLAN, "入口", PLAN_STEP_ORDER, customer=cu, sub=10),
              plan_feat(cu), "入口", "需求录入", "方案速配入口卡", "创建新方案")
add_click(pkey(FI_PLAN, "查看历史", PLAN_STEP_ORDER, customer="老客户", sub=10),
          plan_feat("老客户"), "入口", "查看历史", "方案速配入口卡", "查看历史数据")
add_click(pkey(FI_PLAN, "需求录入", PLAN_STEP_ORDER, customer="老客户", sub=10),
          plan_feat("老客户"), "需求录入", "选品", "需求描述卡", "创建新方案 -> 跳过，按最近订单推荐")
add_click(pkey(FI_PLAN, "需求录入", PLAN_STEP_ORDER, customer="老客户", sub=11),
          plan_feat("老客户"), "需求录入", "选品", "需求描述卡", "跳过")
for cu in ("新客户", "老客户"):
    add_click(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=10),
              plan_feat(cu), "选品", "选品", "方案选品卡", "输入筛选词 -> 筛选")
    add_click(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=11),
              plan_feat(cu), "选品", "方案预览", "方案选品卡", "勾选产品 -> 预览方案")
    if cu == "老客户":
        add_click(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=12),
                  plan_feat(cu), "选品", "选品", "方案选品卡", "补充需求优化推荐")
    add_click(pkey(FI_PLAN, "方案预览", PLAN_STEP_ORDER, customer=cu, sub=10),
              plan_feat(cu), "方案预览", "选品", "方案预览卡", "返回选品")
    add_click(pkey(FI_PLAN, "方案预览", PLAN_STEP_ORDER, customer=cu, sub=11),
              plan_feat(cu), "方案预览", "方案模板", "方案预览卡", "生成方案")
    add_click(pkey(FI_PLAN, "方案模板", PLAN_STEP_ORDER, customer=cu, sub=10),
              plan_feat(cu), "方案模板", "方案成果", "方案模板消息卡", "选标准技术方案 -> 保存方案")
    add_click(pkey(FI_PLAN, "方案模板", PLAN_STEP_ORDER, customer=cu, sub=11),
              plan_feat(cu), "方案模板", "方案成果", "方案模板消息卡", "第1个 -> 保存方案")
    add_click(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=13),
              plan_feat(cu), "选品", "方案成果", "方案选品卡",
              "勾选 -> 改规格 -> 预览方案 -> 生成方案 -> 保存方案")

for br, cu in QUOTE_BRANCHES:
    sk = lambda st, s=10: pkey(FI_QUOTE, st, QUOTE_STEP_ORDER, br, cu, s)
    feat = quote_feat(br, cu)
    add_click(sk("入口"), feat, "入口", "报价来源", "产品报价入口卡", "新建报价")
    add_click(sk("入口", 1), feat, "入口", "查看历史", "产品报价入口卡", "查看历史报价单")
    tgt_after_source = "选择方案" if br == "按方案" else "选品"
    add_click(sk("报价来源"), feat, "报价来源", tgt_after_source, "报价来源卡",
              "按方案报价" if br == "按方案" else "直接选品报价")
    if br == "按方案":
        add_click(sk("选择方案"), feat, "选择方案", "逐项报价", "选择方案卡", "按方案报价 -> 点列表第1行")
    if br == "自选":
        add_click(sk("选品"), feat, "选品", "逐项报价", "报价选品卡", "勾选产品 -> 下一步")
        if cu == "新客户":
            add_click(sk("需求录入"), feat, "报价来源", "需求录入", "报价需求描述卡", "直接选品报价")
    add_click(sk("逐项报价"), feat, "逐项报价", "报价单模板", "逐项报价消息卡", "选择模板")
    add_click(sk("报价单模板"), feat, "报价单模板", "报价成果", "报价单模板消息卡",
              "选标准销售报价单 -> 生成报价单")

for br, cu in ORDER_BRANCHES:
    sk = lambda st, s=10: pkey(FI_ORDER, st, ORDER_STEP_ORDER, br, cu, s)
    feat = order_feat(br, cu)
    tgt = "选择报价单" if br == "按报价单" else "选品"
    add_click(sk("下单来源"), feat, "下单来源", tgt, "下单来源卡",
              "按报价单下单" if br == "按报价单" else "直接选品")
    if br == "按报价单":
        add_click(sk("选择报价单"), feat, "选择报价单", "订单确认", "选择报价单卡",
                  "按报价单下单 -> 点第1行")
    if br == "自选":
        add_click(sk("选品"), feat, "选品", "订单确认", "订单选品卡",
                  "选品 -> 加购 -> 逐项报价 -> 下一步")
    add_click(sk("订单确认"), feat, "订单确认", "订单成功", "订单确认消息卡", "确认下单")

add_click((FI_MISC, 3, -1, -1, 10), plan_feat("老客户"), "方案成果", "报价来源",
          "方案成果卡", "预览电子文档 -> 去报价")
add_click((FI_MISC, 3, -1, -1, 11), quote_feat("按方案", "老客户"), "报价成果", "下单来源",
          "报价单成果卡", "预览电子文档 -> 生成订单")
add_click((FI_MISC, 4, -1, -1, 10), "跨功能切换", "跨功能切换", "跨功能切换",
          "跨功能确认消息卡", "使用当前信息 / 重新走流程")
add_click((FI_MISC, 5, -1, -1, 10), "对话页", "对话页", "选客户", "选客户消息卡",
          "切换客户 -> 搜索华东 -> 点选华东精密")
add_click((FI_MISC, 6, -1, -1, 10), "待跟进", "待跟进", "入口", "待跟进列表卡", "做方案")

# ── 表单填充 ─────────────────────────────────────────────
for cu in ("新客户", "老客户"):
    add_fill(pkey(FI_PLAN, "需求录入", PLAN_STEP_ORDER, customer=cu, sub=20),
             plan_feat(cu), "需求录入", "选品", "需求描述卡", "伺服电机和传动齿轮箱各2台",
             "客户、已在需求描述卡", "需求描述")
    add_fill(pkey(FI_PLAN, "选品", PLAN_STEP_ORDER, customer=cu, sub=20),
             plan_feat(cu), "选品", "选品", "方案选品卡", "筛选 伺服 / 选品 伺服电机",
             "客户、已在选品卡", "筛选词或品名")
    add_fill(pkey(FI_PLAN, "方案预览", PLAN_STEP_ORDER, customer=cu, sub=20),
             plan_feat(cu), "方案预览", "方案预览", "方案预览卡", "伺服电机 数量 3",
             "客户、已在预览卡", "数量")
for br, cu in QUOTE_BRANCHES:
    feat = quote_feat(br, cu)
    add_fill(pkey(FI_QUOTE, "逐项报价", QUOTE_STEP_ORDER, br, cu, 20),
             feat, "逐项报价", "逐项报价", "逐项报价消息卡", "伺服电机 报价 4200 / 打九折",
             "客户、已在逐项报价卡", "本单报价/折扣")
    if br == "自选":
        add_fill(pkey(FI_QUOTE, "选品", QUOTE_STEP_ORDER, br, cu, 20),
                 feat, "选品", "选品", "报价选品卡", "筛选 电机 / 选品 伺服电机",
                 "客户、已在选品卡", "筛选词或品名")
for br, cu in ORDER_BRANCHES:
    feat = order_feat(br, cu)
    add_fill(pkey(FI_ORDER, "逐项报价", ORDER_STEP_ORDER, br, cu, 20),
             feat, "逐项报价", "逐项报价", "逐项报价消息卡", "伺服电机 报价 2100",
             "客户、已在逐项报价卡", "本单报价")

MISC_ROWS.append(row("对话页", "对话页", "对话页", "对话页", "能力说明", "帮助 / 能做什么", "—", "可识别意图"))


def build_rows():
    def sorted_part(rows):
        return [r for _, r in sorted(rows, key=lambda x: x[0])]
    return (COMPLETE_ROWS + sorted_part(PARTIAL_ROWS) + sorted_part(CLICK_ROWS)
            + sorted_part(FILL_ROWS) + MISC_ROWS)


def main():
    rows = build_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "路由指令穷举"
    hf = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font = hf, hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([22, 12, 12, 26, 12, 40, 32, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows) + 1}"

    meta = wb.create_sheet("说明")
    meta["A3"] = "列说明"
    meta["B3"] = (
        "功能名称：三大功能 + 括号变体\n"
        "当前步骤：处理本句指令时用户所处流程步骤（或对话页）\n"
        "目标步骤：系统应推进到的步骤（缺槽时为首缺子步骤；点击为下一步）\n"
        "目标子步骤/界面名称：目标步骤对应的消息流卡片\n"
        "已提供 / 待提供：槽位分列"
    )
    meta["B3"].alignment = Alignment(wrap_text=True)
    meta.column_dimensions["B"].width = 90
    meta["A5"] = "统计"
    meta["B5"] = (f"完整{len(COMPLETE_ROWS)} + 部分{len(PARTIAL_ROWS)} + "
                  f"模拟点击{len(CLICK_ROWS)} + 表单填充{len(FILL_ROWS)} = {len(rows)}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {len(rows)} rows -> {OUT}")


if __name__ == "__main__":
    main()
