#!/usr/bin/env python3
"""从 QA对 Excel 生成 intent-qa.generated.js（SSOT → 运行时词典）。"""
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "QA对-主功能口语映射-v1.5.0.xlsx"
OUT = ROOT / "js" / "intent-qa.generated.js"

FEATURE_IDS = {
    "今日待跟": "followup",
    "新增客户": "customer-create",
    "方案速配": "plan",
    "产品报价": "quote",
    "交期评审": "delivery",
    "确认下单": "order",
    "复制订单": "copy",
    "订单变更": "change",
    "订单进度": "progress",
    "产能分析": "capacity",
    "库存查询": "inventory",
    "业务分析": "biz-analysis",
    "回款分析": "payment",
    "切换客户": "switch-customer",
}


def normalize(text: str) -> str:
    t = (text or "").strip()
    t = re.sub(r"\s+", "", t)
    t = re.sub(r"[，,。.；;！!？?、]", "", t)
    return t


def normalize_pair_type(raw) -> str:
    s = (raw or "").strip()
    if s in ("knowledge", "知识"):
        return "knowledge"
    return "function"


def parse_header(header: list[str]) -> dict:
    type_idx = header.index("类型") if "类型" in header else -1
    note_idx = header.index("备注") if "备注" in header else -1
    if type_idx >= 0:
        return {"type_idx": type_idx, "note_idx": note_idx if note_idx >= 0 else 4}
    has_group = "分组" in header
    return {"type_idx": -1, "note_idx": 4 if has_group else 3}


def load_pairs() -> list[dict]:
    if not XLSX.exists():
        raise SystemExit(f"缺少 Excel：{XLSX}")

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if "QA对" not in wb.sheetnames:
        raise SystemExit("Excel 缺少 sheet「QA对」")
    ws = wb["QA对"]
    header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())]
    col = parse_header(header)
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    pairs: list[dict] = []
    errors: list[str] = []
    seen_row: set[tuple[str, str, str]] = set()

    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        _id, q, a = (row + (None,) * 3)[:3]
        q = (q or "").strip()
        a = (a or "").strip()
        pair_type = normalize_pair_type(row[col["type_idx"]] if col["type_idx"] >= 0 else "")
        note = (row[col["note_idx"]] or "").strip() if col["note_idx"] >= 0 else ""

        if not q or not a:
            errors.append(f"第 {i} 行 Q/A 为空")
            continue
        if pair_type == "function" and a not in FEATURE_IDS:
            errors.append(f"第 {i} 行未知功能 A={a!r}")
            continue
        key = (q, a, pair_type)
        if key in seen_row:
            continue
        seen_row.add(key)
        item = {
            "q": q,
            "a": a,
            "pairType": pair_type,
            "note": note,
        }
        if pair_type == "function":
            item["aId"] = FEATURE_IDS[a]
        pairs.append(item)

    if errors:
        raise SystemExit("校验失败：\n" + "\n".join(errors))
    if not pairs:
        raise SystemExit("QA对 sheet 无有效数据")

    return pairs


def build_multi_q_index(pairs: list[dict]) -> list[dict]:
    by_q: dict[str, list[str]] = {}
    for p in pairs:
        if p.get("pairType", "function") != "function":
            continue
        by_q.setdefault(p["q"], []).append(p["a"])
    multi = []
    for q, als in sorted(by_q.items()):
        uniq = []
        for a in als:
            if a not in uniq:
                uniq.append(a)
        if len(uniq) >= 2:
            multi.append({"q": q, "a": uniq})
    return multi


def emit_js(pairs: list[dict]) -> str:
    multi = build_multi_q_index(pairs)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    payload = {
        "version": "v1.5.0",
        "source": XLSX.name,
        "generatedAt": now,
        "algorithm": "C",
        "pairCount": len(pairs),
        "uniqueQCount": len({p["q"] for p in pairs}),
        "pairs": pairs,
        "multiQ": multi,
        "featureIds": FEATURE_IDS,
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        "/** 自动生成 · 勿手改。源："
        + XLSX.name
        + " · 运行：python3 scripts/build-intent-qa.py */\n"
        "(function () {\n"
        "  window.IntentQa = "
        + body
        + ";\n"
        "})();\n"
    )


def main() -> None:
    pairs = load_pairs()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(emit_js(pairs), encoding="utf-8")
    multi_n = len(build_multi_q_index(pairs))
    fn_n = sum(1 for p in pairs if p.get("pairType", "function") == "function")
    kn_n = len(pairs) - fn_n
    print(f"✓ 写入 {OUT.relative_to(ROOT)}")
    print(f"  pairs={len(pairs)} function={fn_n} knowledge={kn_n} multi_q={multi_n}")


if __name__ == "__main__":
    main()
