# -*- coding: utf-8 -*-
"""生成 v1.5.0 各功能子卡片清单 Excel。"""
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REPO = ROOT.parent


def load_annotation_names():
    names = {}
    for rel in [
        "v1.4.0/js/annotation-spec-data.js",
        "v1.5.0/js/annotation-spec-data.js",
    ]:
        path = REPO / rel
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8")
        for m in re.finditer(r"'([a-z0-9-]+)'\s*:\s*\{[^}]*?name:\s*'([^']+)'", text, re.S):
            names[m.group(1)] = m.group(2)
    return names


def n(spec_id, fallback, ann):
    return ann.get(spec_id, fallback)


def rows(ann):
    """返回 list[tuple]: feature_id, feature_name, step, spec_id, card_name, card_type, when]"""
    A = ann
    data = []

    def add(fid, fname, items):
        for step, spec_id, card_name, ctype, when in items:
            data.append((fid, fname, step, spec_id, card_name, ctype, when))

    add("common", "公共 / 首屏 / 跨功能", [
        ("—", "chat-welcome", "首屏 · 欢迎区", "首屏块", "登录后首条助手消息"),
        ("—", "chat-welcome-features", "首屏 · 快捷功能格", "首屏块", "欢迎区下方功能 chip"),
        ("—", "chat-recent", "首屏 · 最近访问", "首屏块", "待跟进摘要下方"),
        ("—", "card-followup-summary", "今日待跟进 · 摘要", "首屏块", "首屏内可展开"),
        ("—", "card-followup-list", "今日待跟进 · 列表", "流程卡", "点摘要/技能/话术展开"),
        ("—", "card-template-followup", "微信模板消息卡", "站外演示", "左侧微信面板"),
        ("—", "card-customer-detail", n("card-customer-detail", "企业详情卡", A), "流程卡", "待跟进选客 / 新增客户提交后"),
        ("—", "card-next-step", n("card-next-step", "下一步引导卡", A), "流程卡", "选客或新增客户提交后"),
        ("—", "sheet-customer", n("sheet-customer", "选客户卡", A), "流程卡", "切换客户 / 缺客户引导 / 回款指定客户"),
        ("—", "card-customer-prompt", n("card-customer-prompt", "缺客户 · 引导卡", A), "流程卡", "需客户技能未选客户时"),
        ("—", "sheet-followup", n("sheet-followup", "写跟进表单卡", A), "表单卡", "下一步点写跟进 / 写跟进技能"),
        ("—", "sheet-enterprise", "选企业卡", "流程卡", "多企业切换（演示）"),
        ("—", "modal-skill-switch", n("modal-skill-switch", "功能切换确认弹窗", A), "弹窗", "切换技能且已选客户时"),
        ("—", "modal-pdf", n("modal-pdf", "PDF 预览", A), "弹窗", "方案/报价预览 PDF"),
        ("—", "modal-logout", "退出登录确认", "弹窗", "顶栏退出"),
        ("—", "modal-cross-skill", "跨技能切换确认卡", "流程卡", "对话内切换技能时"),
        ("—", "chat-bubble-user", "用户消息气泡", "对话块", "用户发送话术后"),
        ("—", "sheet-create-region", "选择地区弹窗", "弹窗", "新增客户表单点地区"),
        ("—", "card-region-picker", n("card-region-picker", "地区选择器", A), "弹窗子卡", "地区弹窗内容区"),
        ("—", "btn-customer-create", n("btn-customer-create", "选客户卡 · 新增客户入口", A), "入口按钮", "选客户卡底部"),
        ("—", "btn-customer-create-skill", n("btn-customer-create-skill", "技能条 · 新增客户", A), "入口按钮", "底栏技能条"),
    ])

    add("followup", "今日待跟", [
        ("1", "card-followup-summary", "今日待跟进 · 摘要", "入口", "首屏或点技能"),
        ("2", "card-followup-list", "今日待跟进 · 列表", "列表卡", "展开待跟进"),
        ("3", "card-customer-detail", "企业详情卡", "详情卡", "点选某客户"),
        ("4", "card-next-step", "下一步引导卡", "引导卡", "选客后"),
        ("5", "sheet-followup", "写跟进表单卡", "表单卡", "点「写跟进」"),
    ])

    add("customer-create", "新增客户", [
        ("1", "card-customer-create", n("card-customer-create", "新增客户表单卡", A), "表单卡", "技能条/选客户卡/话术入口"),
        ("2", "card-region-picker", "地区选择器", "弹窗", "表单内点选地区"),
        ("3", "card-customer-detail", "企业详情卡", "详情卡", "提交成功后"),
        ("4", "card-next-step", "下一步引导卡", "引导卡", "提交成功后"),
    ])

    add("plan", "方案速配", [
        ("1", "card-plan-entry", n("card-plan-entry", "方案速配 · 入口卡", A), "入口卡", "技能条/欢迎区"),
        ("2", "card-plan-demand", n("card-plan-demand", "需求引导卡", A), "引导卡", "点「创建新方案」"),
        ("3", "card-plan-pick", n("card-plan-pick", "方案选品卡", A), "选品卡", "确认需求或跳过"),
        ("4", "card-plan-preview", n("card-plan-preview", "方案预览卡", A), "预览卡", "选品后点预览"),
        ("5", "sheet-plan-template", n("sheet-plan-template", "方案模板选择卡", A), "表单卡", "预览后保存方案"),
        ("6", "card-scheme", n("card-scheme", "方案卡", A), "结果卡", "保存成功 / 点历史方案"),
        ("—", "card-scheme-pick", n("card-scheme-pick", "历史方案列表", A), "列表卡", "入口卡点查看历史"),
    ])

    add("quote", "产品报价", [
        ("1", "card-quote-entry", n("card-quote-entry", "产品报价 · 入口卡", A), "入口卡", "技能条/欢迎区"),
        ("2", "card-quote-demand", n("card-quote-demand", "需求引导卡（报价）", A), "引导卡", "创建新报价"),
        ("3", "card-quote-source", n("card-quote-source", "报价来源卡", A), "分支卡", "选择新建/历史/从方案"),
        ("4", "card-scheme-pick", "历史方案列表", "列表卡", "来源选「从方案」"),
        ("5", "card-quote-pick", n("card-quote-pick", "报价选品卡", A), "选品卡", "新建报价路径"),
        ("6", "card-quote-cart", n("card-quote-cart", "报价选品确认卡", A), "确认卡", "选品后确认"),
        ("7", "sheet-quote-setup", n("sheet-quote-setup", "逐项报价卡", A), "表单卡", "确认后填单价"),
        ("8", "sheet-quote-template", n("sheet-quote-template", "报价单模板选择卡", A), "表单卡", "逐项报价后生成"),
        ("9", "card-quote", n("card-quote", "报价单卡", A), "结果卡", "生成报价单成功"),
        ("—", "card-quote-select", n("card-quote-select", "历史报价单列表", A), "列表卡", "入口点查看历史"),
        ("—", "card-scheme", "方案卡", "结果卡", "从方案带入后展示"),
    ])

    add("delivery", "交期评审", [
        ("1", "card-delivery-entry", n("card-delivery-entry", "交期评审 · 入口卡", A), "入口卡", "技能条/话术"),
        ("2", "card-delivery-source", n("card-delivery-source", "交期评审 · 选择来源", A), "分支卡", "点评估交期"),
        ("3", "card-delivery-scheme-pick", n("card-delivery-scheme-pick", "交期 · 选方案", A), "列表卡", "来源选方案"),
        ("4", "card-delivery-quote-pick", n("card-delivery-quote-pick", "交期 · 选报价单", A), "列表卡", "来源选报价单"),
        ("5", "card-delivery-order-pick", n("card-delivery-order-pick", "交期 · 选订单", A), "列表卡", "来源选订单"),
        ("6", "card-delivery-demand", n("card-delivery-demand", "交期 · 需求引导", A), "引导卡", "按订单且无行时"),
        ("7", "sheet-delivery", n("sheet-delivery", "交期评审 · 表单", A), "表单卡", "选定单据后填交期"),
        ("8", "card-delivery-line-pick", n("card-delivery-line-pick", "交期评审 · 选明细行", A), "选品卡", "表单内勾选评审行"),
        ("9", "card-delivery", n("card-delivery", "交期评审 · 结果卡", A), "结果卡", "提交评审后"),
    ])

    add("order", "确认下单", [
        ("1", "card-order-entry", n("card-order-entry", "确认下单 · 入口卡", A), "入口卡", "技能条/欢迎区"),
        ("2", "card-order-source", n("card-order-source", "下单来源卡", A), "分支卡", "点确认下单"),
        ("3", "card-quote-pick", "报价选品卡", "选品卡", "来源「按报价」新建路径"),
        ("4", "card-quote-select", "历史报价单列表", "列表卡", "来源选历史报价"),
        ("5", "card-order-pick", n("card-order-pick", "订单选品卡", A), "选品卡", "从报价单/购物车确认"),
        ("6", "card-order-cart", n("card-order-cart", "订单购物车卡", A), "确认卡", "选品后确认数量规格"),
        ("7", "sheet-order", n("sheet-order", "确认下单卡", A), "表单卡", "填交货日期等"),
        ("8", "card-order-success", n("card-order-success", "订单提交成功", A), "结果卡", "提交订单后"),
        ("—", "card-order", n("card-order", "订单卡", A), "结果卡", "查看历史订单详情"),
        ("—", "card-order-select", n("card-order-select", "历史订单列表", A), "列表卡", "入口查看历史"),
    ])

    add("copy", "复制订单", [
        ("1", "card-copy-demand", n("card-copy-demand", "复制订单 · 需求筛选", A), "引导卡", "技能「复制订单」"),
        ("2", "card-order-pick", "历史订单列表（复制）", "列表卡", "确认筛选后"),
        ("3", "card-order-copy-line-pick", n("card-order-copy-line-pick", "复制订单 · 勾选货品", A), "选品卡", "选订单后勾选行"),
        ("4", "sheet-order", "确认下单卡", "表单卡", "勾选后确认下单"),
        ("5", "card-order-success", "订单提交成功", "结果卡", "提交后"),
        ("—", "card-order-copy", n("card-order-copy", "复制订单 · 明细确认", A), "备用卡", "代码保留，主链路已不经此卡"),
    ])

    add("change", "订单变更", [
        ("1", "card-order-pick", "历史订单列表（变更）", "列表卡", "技能「订单变更」"),
        ("2", "card-change-confirm", n("card-change-confirm", "订单变更 · 确认变更", A), "表单卡", "选订单后"),
        ("3", "card-change-success", n("card-change-success", "变更已提交", A), "结果卡", "提交变更后"),
    ])

    add("progress", "订单进度", [
        ("1", "card-progress-demand", n("card-progress-demand", "订单进度 · 需求筛选", A), "引导卡", "技能「订单进度」"),
        ("2", "card-order-pick", "历史订单列表（进度）", "列表卡", "确认筛选后"),
        ("3", "card-order-progress-detail", n("card-order-progress-detail", "订单进度 · 详情", A), "结果卡", "选订单后"),
    ])

    add("capacity", "产能分析", [
        ("1", "card-capacity", n("card-capacity", "产能分析", A), "结果卡", "技能直达 / 话术"),
        ("2", "card-capacity-block-detail", n("card-capacity-block-detail", "产能 · 订单详情 Tab", A), "子面板", "结果卡内点占用块"),
    ])

    add("inventory", "库存查询", [
        ("1", "card-inventory-filter", n("card-inventory-filter", "库存查询 · 筛选卡", A), "筛选卡", "技能进入"),
        ("2", "card-inventory", n("card-inventory", "库存查询 · 结果", A), "结果卡", "筛选后展示"),
    ])

    add("biz-analysis", "业务分析", [
        ("1", "card-biz-analysis-insight", n("card-biz-analysis-insight", "业务分析 · 一句话结论", A), "摘要卡", "技能进入"),
        ("2", "card-biz-date-range-picker", n("card-biz-date-range-picker", "业务分析 · 时间范围", A), "筛选卡", "切换统计区间"),
        ("3", "card-biz-analysis", n("card-biz-analysis", "业务分析", A), "结果卡", "展示图表与明细"),
    ])

    add("payment", "回款分析", [
        ("1", "card-payment-year", "回款分析 · 选年份", "分步卡 Step1", "技能进入 / 缺年份"),
        ("2", "card-payment-scope", "回款分析 · 选范围", "分步卡 Step2", "选定年份后"),
        ("3", "sheet-customer", "选客户卡", "流程卡", "范围选「指定客户」"),
        ("4", "card-payment", n("card-payment", "回款分析 · 结果", A), "结果卡", "确认范围后"),
    ])

    return data


def build_workbook(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "功能子卡片清单"

    headers = [
        "功能ID",
        "功能名称",
        "流程序号",
        "子卡片ID",
        "子卡片名称",
        "卡片类型",
        "典型出现时机",
    ]
    header_fill = PatternFill("solid", fgColor="108482")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [14, 16, 10, 28, 28, 14, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # 功能索引 sheet
    idx = wb.create_sheet("功能索引")
    idx.append(["功能ID", "功能名称", "子卡片数量"])
    idx[1][0].fill = header_fill
    idx[1][1].fill = header_fill
    idx[1][2].fill = header_fill
    for c in range(1, 4):
        idx[1][c - 1].font = header_font

    from collections import OrderedDict
    counts = OrderedDict()
    for fid, fname, *_ in data:
        if fid == "common":
            continue
        key = (fid, fname)
        counts[key] = counts.get(key, 0) + 1
    for (fid, fname), cnt in counts.items():
        idx.append([fid, fname, cnt])
    idx.column_dimensions["A"].width = 16
    idx.column_dimensions["B"].width = 18
    idx.column_dimensions["C"].width = 12

    # 说明 sheet
    info = wb.create_sheet("说明")
    info["A1"] = "文档说明"
    info["A1"].font = Font(bold=True, size=12)
    lines = [
        "版本：v1.5.0 H5 原型",
        "数据来源：v1.5.0/js/skills.js、app.js、v155-features.js 及标注文档",
        "",
        "字段说明：",
        "· 流程序号：主流程顺序；「—」表示分支/复用/非严格顺序",
        "· 子卡片ID：对应 HTML data-spec-id，用于标注与验收对照",
        "· 卡片类型：入口卡 / 引导卡 / 选品卡 / 表单卡 / 结果卡 / 弹窗 等",
        "",
        "备注：",
        "· 公共区卡片可被多个功能复用（如选客户卡、企业详情卡）",
        "· 跨技能卡片（如 card-quote-pick）在所属主流程与复用功能下均有列出",
    ]
    for i, line in enumerate(lines, 2):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 72

    return wb


def main():
    ann = load_annotation_names()
    data = rows(ann)
    out = REPO / ".output" / "功能子卡片清单-v1.5.0.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(data)
    wb.save(out)
    print(str(out))
    print(f"rows={len(data)}")


if __name__ == "__main__":
    main()
